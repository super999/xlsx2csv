#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Time    : 2025/2/11 16:20
# @Author  : ChenXiaWen
# @File    : easy_xlsx_to_csv.py

import logging
import argparse
import openpyxl
import csv
from pathlib import Path
from glob import glob
import sys


def convert_xlsx_to_csv(input_path, output_dir, args):
    """将单个XLSX文件转换为CSV"""
    try:
        if args.verbose:
            logging.info(f"正在读取Excel文件: {input_path}")

        # 加载工作簿
        wb = openpyxl.load_workbook(input_path, data_only=True)

        # 确定要处理的工作表,
        first_sheet_name = wb.sheetnames[0]
        sheets = wb.sheetnames if args.all_sheets else [first_sheet_name]

        for sheet_name in sheets:
            try:
                # 获取工作表
                ws = wb[sheet_name]

                # 生成输出文件名
                base_name = input_path.stem
                sheet_suffix = f"_{sheet_name}" if args.all_sheets else ""
                output_file = output_dir / f"{base_name}{sheet_suffix}.csv"

                # 检查文件是否存在，且设置不允许覆盖
                if output_file.exists():
                    if args.no_force:
                        logging.warning(f"文件已存在: {output_file}, 文件跳过")
                        continue
                    else:
                        if args.verbose:
                            logging.warning(f"文件已存在，正在覆盖: {output_file}")

                # 写入CSV
                with open(output_file, mode='w', newline='', encoding=args.encoding) as f:
                    writer = csv.writer(f, delimiter=args.delimiter)

                    k = ws.rows
                    for row in ws.iter_rows(values_only=True):
                        # 将所有布尔值转换为大写字符串，其余值保持原样
                        formatted_row = [str(cell).upper() if isinstance(cell, bool) else cell for cell in row]
                        writer.writerow(formatted_row)

                if args.chinese:
                    logging.info(f"成功创建文件: {output_file}")
                else:
                    logging.info(f"Generated file: {output_file}")

            except Exception as e:
                logging.exception(f"处理工作表 {sheet_name} 时出错")

    except Exception as e:
        print(f"错误: {str(e)}", file=sys.stderr)
        sys.exit(1)


def init_logging(args):
    if args.logformat:
        logging.basicConfig(level=logging.INFO, format=args.logformat)
    else:
        logging.basicConfig(level=logging.INFO, format='%(message)s')


def main():
    parser = argparse.ArgumentParser(
        description='将XLSX文件转换为CSV格式',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument('input', help='输入的XLSX文件路径、文件夹路径或通配符模式')
    parser.add_argument('-o', '--output', help='输出目录', default='./')
    parser.add_argument('-s', '--sheet', help='工作表名称或索引（从0开始）', default=0)
    parser.add_argument('-d', '--delimiter', help='CSV文件的分隔符', default=',')
    parser.add_argument('-e', '--encoding', help='输出文件的编码格式', default='utf-8-sig')
    parser.add_argument('--no-header', action='store_true', help='不包含标题行')
    parser.add_argument('--index', action='store_true', help='包含索引列')
    parser.add_argument('-nf', '--no-force', action='store_true', help='不覆盖现有文件')
    parser.add_argument('-v', '--verbose', action='store_true', help='显示详细输出信息')
    parser.add_argument('-c', '--chinese', action='store_true', help='输出中文提示')
    parser.add_argument('-a', '--all-sheets', action='store_true', help='转换工作簿中的所有工作表')
    # logging format 的模版 : %(asctime)s - %(name)s - %(levelname)s - %(message)s
    parser.add_argument('-f', '--logformat', help='logging format', default='%(message)s')

    args = parser.parse_args()

    init_logging(args)

    # if args.verbose:
    logging.info(f"ExcelToCsv Start...")

    try:
        # 处理输入路径
        input_path = Path(args.input).resolve()
        if not input_path.exists():
            raise FileNotFoundError(f"输入路径不存在: {input_path}")

        # 处理输出目录
        output_dir = Path(args.output).resolve()
        output_dir.mkdir(parents=True, exist_ok=True)

        # 获取所有XLSX文件
        if input_path.is_dir():
            xlsx_files = list(input_path.glob("*.xlsx"))
        elif "*" in str(input_path):  # 处理通配符
            xlsx_files = [Path(f) for f in glob(str(input_path))]
        else:
            xlsx_files = [input_path]

        if not xlsx_files:
            raise ValueError("未找到任何XLSX文件")

        # 如果 xlsx_files 带 ~$ 开头的临时文件，去掉
        xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

        if args.verbose:
            logging.info(f"找到 {len(xlsx_files)} 个XLSX文件")

        # 处理每个文件
        for xlsx_file in xlsx_files:
            if args.verbose:
                logging.info(f"开始处理文件: {xlsx_file}")

            # 如果输入是文件夹，保持输出目录结构
            if input_path.is_dir():
                relative_path = xlsx_file.relative_to(input_path).parent
                current_output_dir = output_dir / relative_path
                current_output_dir.mkdir(parents=True, exist_ok=True)
            else:
                current_output_dir = output_dir

            convert_xlsx_to_csv(xlsx_file, current_output_dir, args)

        # if args.verbose:
        logging.info("ExcelToCsv End. 转换完成")

    except Exception as e:
        print(f"错误: {str(e)}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
